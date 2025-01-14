#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import asyncio
from difflib import SequenceMatcher
import openpyxl
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import nest_asyncio
import sys
import os

# Set up Playwright environment for packaged executables
if getattr(sys, 'frozen', False):  # Check if running as a PyInstaller executable
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "0"  # Use bundled binaries
    from playwright._impl._path_utils import get_asar_path
    asar_path = os.path.join(sys._MEIPASS, "playwright")  # Locate Playwright package
    get_asar_path().overwrite(asar_path)

nest_asyncio.apply()


async def extract_links_from_sitemap(sitemap_url):
    links = []
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context()
            page = await context.new_page()
            await page.goto(sitemap_url, wait_until="domcontentloaded", timeout=30000)
            content = await page.content()
            soup = BeautifulSoup(content, features="xml")
            links = [loc.text.strip() for loc in soup.find_all("loc")]
            await browser.close()
    except Exception as e:
        print(f"Error fetching sitemap: {e}")
    return links


async def check_url_status(url, page, retries=3):
    for attempt in range(retries):
        try:
            response = await page.goto(url, wait_until="domcontentloaded", timeout=15000)
            if response:
                status_code = response.status
                print(f"Checked URL: {url} | Status Code: {status_code}")

                if status_code == 404:
                    return status_code, True

                content = await page.content()
                if "Page Not Found" in content or "404" in content:
                    return status_code, True

                return status_code, False  # Not broken

        except Exception as e:
            print(f"Attempt {attempt + 1} failed for URL: {url} | Error: {e}")
            await asyncio.sleep(2)

    print(f"Failed to fetch URL after {retries} attempts: {url}")
    return None, True


async def process_urls(broken_urls, sitemap_links, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "URL Mapping"
    sheet.append(["Broken URL", "Suggested URL", "Similarity Score", "Status Code"])

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        for broken_url in broken_urls:
            print(f"Processing: {broken_url}")
            status_code, is_broken = await check_url_status(broken_url, page)

            if status_code is None:
                sheet.append([broken_url, "Error - No response", "Error", "Error"])
                continue

            if not is_broken:
                sheet.append([broken_url, "N/A - page is not broken", "N/A", status_code])
                continue

            best_match = None
            highest_similarity = 0
            for link in sitemap_links:
                similarity = SequenceMatcher(None, broken_url, link).ratio()
                if similarity > highest_similarity:
                    best_match = link
                    highest_similarity = similarity

            sheet.append([broken_url, best_match, round(highest_similarity, 2), status_code])

        await browser.close()

    workbook.save(output_file)
    print(f"Mapping results saved to {output_file}")


async def main():
    sitemap_url = input("Enter the XML Sitemap URL: ")
    broken_urls_file = "broken urls.xlsx"
    output_file = "url_mapping.xlsx"

    print(f"Fetching URLs from sitemap: {sitemap_url}")
    sitemap_links = await extract_links_from_sitemap(sitemap_url)

    broken_workbook = openpyxl.load_workbook(broken_urls_file)
    broken_sheet = broken_workbook.active
    broken_urls = [
        row[0].value for row in broken_sheet.iter_rows(min_row=2, min_col=1, max_col=1)
        if row[0].value
    ]

    print("Processing broken URLs...")
    await process_urls(broken_urls, sitemap_links, output_file)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("Script interrupted by user.")


# In[ ]:




